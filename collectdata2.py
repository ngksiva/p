import asyncio
from concurrent.futures import ThreadPoolExecutor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from telethon import TelegramClient
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pytz import timezone
import time

# Путь к драйверу Chrome
driver_path = "/opt/homebrew/bin/chromedriver"  # путь к chromedriver

# Файлы
input_file = "channels.txt"
output_excel = "channels.xlsx"
output_html = "channels.html"

# Получаем API ID и HASH из переменных окружения
api_id = os.getenv('TELEGRAM_API_ID')
api_hash = os.getenv('TELEGRAM_API_HASH')

# Проверяем, что значения были загружены
if not api_id or not api_hash:
    print("API ID или API HASH не найдены в переменных окружения. Убедитесь, что они заданы.")
    exit(1)

# Настройка веб-драйвера
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # Запуск без открытия окна браузера (опционально)


def init_driver():
    service = webdriver.chrome.service.Service(driver_path)
    return webdriver.Chrome(service=service, options=options)


# Функция для извлечения даты последнего сообщения через Telethon
async def get_last_message_date_telethon(channel_username):
    async with TelegramClient('session_name', api_id, api_hash) as client:
        try:
            messages = await client.get_messages(channel_username, limit=1)
            if messages:
                return messages[0].date
            else:
                return None
        except Exception as e:
            print(f"Ошибка при запросе данных через API для {channel_username}: {e}")
            return None


# Функция для извлечения данных через Selenium
def get_channel_data(url):
    driver = None
    try:
        driver = init_driver()
        driver.get(url)
        time.sleep(5)  # Задержка для загрузки страницы

        # Проверяем существование канала/чата
        try:
            channel_name = driver.find_element(By.CLASS_NAME, "tgme_page_title").text
        except NoSuchElementException:
            print(f"Канал или чат не существует: {url}")
            return None  # Возвращаем None, если канал не найден

        # Извлечение числа подписчиков/участников
        subscribers_text = driver.find_element(By.CLASS_NAME, "tgme_page_extra").text
        description_html = driver.find_element(By.CLASS_NAME, "tgme_page_description").get_attribute("outerHTML") if driver.find_elements(By.CLASS_NAME, "tgme_page_description") else "Нет описания"

        # Удаление тегов <div class="tgme_page_description" dir="auto"> и </div>
        description_text = description_html.replace('<div class="tgme_page_description" dir="auto">', "").replace("</div>", "")

        # Определение типа объекта (канал или чат)
        if "subscribers" in subscribers_text:
            obj_type = "канал"
            subscribers_count = ''.join(subscribers_text.split()[:-1])  # Собираем число подписчиков до слова "subscribers"
        elif "members" in subscribers_text:
            obj_type = "чат"
            try:
                members_part = subscribers_text.split("members")[0].strip()  # Извлекаем всё перед "members"
                subscribers_count = members_part.replace(" ", "")  # Удаляем пробелы внутри числа
            except Exception:
                subscribers_count = None
        else:
            subscribers_count = None

        if subscribers_count:
            subscribers_count = int(subscribers_count)  # Преобразуем строку в число

        # Формируем ссылку на tgstat
        tgstat_url = f"https://tgstat.ru/channel/@{url.split('/')[-1]}/stat"

        return url, channel_name, subscribers_count, obj_type, description_text, tgstat_url
    except Exception as e:
        print(f"Ошибка при обработке {url}: {e}")
        return None
    finally:
        if driver:
            driver.quit()


async def process_urls(urls):
    results = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        loop = asyncio.get_event_loop()
        tasks = [
            loop.run_in_executor(executor, get_channel_data, url)
            for url in urls
        ]
        for completed_task in asyncio.as_completed(tasks):
            result = await completed_task
            if result:
                telethon_date = await get_last_message_date_telethon(result[0].split('/')[-1])
                if telethon_date:
                    moscow_tz = timezone("Europe/Moscow")
                    telethon_date_moscow = telethon_date.astimezone(moscow_tz)

                    # Форматирование даты
                    formatted_date = telethon_date_moscow.strftime('%d.%m.%Y')
                    print(f"Обработан канал/чат: {result[1]}, Подписчики: {result[2]}, Дата: {formatted_date}, URL: {result[0]}")

                    result = result + (telethon_date_moscow,)
                else:
                    print(f"Обработан канал/чат: {result[1]}, Подписчики: {result[2]}, Дата: Не удалось получить, URL: {result[0]}")
                    result = result + (None,)

                results.append(result)
    return results


def create_html(results):
    html_content = """
    <html>
    <head>
        <title>Telegram Channels</title>
        <style>
            table {
                width: 100%;
                border-collapse: collapse;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
            }
            th {
                background-color: #f2f2f2;
                text-align: left;
            }
            .inactive {
                color: gray;
            }
        </style>
    </head>
    <body>
        <h1>Список каналов и чатов Telegram</h1>
        <table>
            <tr>
                <th>#</th>
                <th>Название</th>
                <th>Подписчики</th>
                <th>Описание</th>
                <th>Дата последнего сообщения</th>
            </tr>
    """
    for index, (url, channel_name, subscribers_count, obj_type, description_text, tgstat_url, last_message_date) in enumerate(results, start=1):
        # Проверяем, активен ли канал/чат
        inactive = False
        if last_message_date:
            one_month_ago = datetime.now() - timedelta(days=30)
            if last_message_date.replace(tzinfo=None) < one_month_ago:
                inactive = True

        # Форматируем дату
        last_message_date_text = last_message_date.strftime('%d.%m.%Y') if last_message_date else "Неизвестно"

        # Добавляем класс inactive для неактивных
        row_class = "inactive" if inactive else ""
        channel_display_name = f"{obj_type} <a href='{url}'>{channel_name}</a>"
        if inactive:
            channel_display_name += " (неактивен)"

        html_content += f"""
            <tr class="{row_class}">
                <td>{index}</td>
                <td>
                    {channel_display_name}<br>
                    <a href="{tgstat_url}">tgstat</a>
                </td>
                <td>{subscribers_count}</td>
                <td>{description_text}</td>
                <td>{last_message_date_text}</td>
            </tr>
        """
    html_content += """
        </table>
    </body>
    </html>
    """
    with open(output_html, "w", encoding="utf-8") as file:
        file.write(html_content)


async def main():
    with open(input_file, 'r', encoding='utf-8') as file:
        urls = [line.strip() for line in file if line.strip()]

    print("Начинаю обработку URL...")
    results = await process_urls(urls)
    print("Обработка завершена.")

    results = sorted(results, key=lambda x: x[2] if x[2] is not None else 0, reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Telegram Channels"

    ws.append(["#", "Тип", "Название", "Подписчики", "Описание", "URL", "TGStat URL", "Дата последнего сообщения"])
    for col in ws[1]:
        col.font = Font(bold=True)

    for index, (url, channel_name, subscribers_count, obj_type, description_text, tgstat_url, last_message_date) in enumerate(results, start=1):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=obj_type)
        ws.cell(row=row, column=3, value=channel_name)
        ws.cell(row=row, column=4, value=subscribers_count)
        ws.cell(row=row, column=5, value=description_text)
        ws.cell(row=row, column=6, value=url)
        ws.cell(row=row, column=7, value=tgstat_url)
        if last_message_date:
            date_without_tz = last_message_date.replace(tzinfo=None)
            formatted_date = date_without_tz.strftime('%d.%m.%Y')
            ws.cell(row=row, column=8, value=formatted_date)

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_excel)
    print(f"Данные успешно сохранены в {output_excel}")

    create_html(results)
    print(f"HTML-файл успешно сохранен в {output_html}")


if __name__ == "__main__":
    asyncio.run(main())
